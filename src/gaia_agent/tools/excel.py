from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook
from smolagents import Tool

from gaia_agent.llm.service import LLMService
from gaia_agent.tools.path_utils import (
    is_placeholder_path,
    resolve_file,
)


EXCEL_EXTENSIONS = {
    ".xlsx",
    ".xlsm",
}

DELIMITED_EXTENSIONS = {
    ".csv",
    ".tsv",
}


class AnalyzeExcelTool(Tool):

    name = "analyze_excel"

    description = (
        "Analyze an Excel workbook or tabular spreadsheet "
        "according to a user's question. Supports XLSX, "
        "XLSM, CSV and TSV files."
    )

    inputs = {
        "file_path": {
            "type": "string",
            "description": (
                "Path to the spreadsheet relative to the "
                "allowed base directory or filename."
            ),
        },
        "question": {
            "type": "string",
            "description": (
                "Question that should be answered using "
                "the spreadsheet data."
            ),
        },
    }

    output_type = "string"

    MAX_ROWS = 5000
    MAX_COLUMNS = 100
    MAX_CHARS = 100_000

    def __init__(
        self,
        llm_service: LLMService,
        base_dir: str = ".",
    ) -> None:
        super().__init__()

        if llm_service is None:
            raise ValueError(
                "AnalyzeExcelTool requires an LLMService."
            )

        self.llm_service = llm_service
        self.base_dir = Path(base_dir).resolve()

    def _read_excel(
        self,
        path: Path,
    ) -> str:
        workbook = load_workbook(
            filename=path,
            data_only=True,
            read_only=True,
        )

        try:
            output: list[str] = []

            for sheet in workbook.worksheets:
                output.append(
                    f"Sheet: {sheet.title}"
                )

                row_count = 0

                for row in sheet.iter_rows(
                    values_only=True
                ):
                    if row_count >= self.MAX_ROWS:
                        output.append(
                            "... [spreadsheet truncated "
                            "because it exceeded the row limit] ..."
                        )
                        break

                    values = [
                        ""
                        if value is None
                        else str(value)
                        for value in row[: self.MAX_COLUMNS]
                    ]

                    if not any(
                        value.strip()
                        for value in values
                    ):
                        continue

                    output.append(
                        " | ".join(values)
                    )

                    row_count += 1

                output.append("")

            return self._truncate(
                "\n".join(output)
            )

        finally:
            workbook.close()

    def _read_delimited(
        self,
        path: Path,
    ) -> str:
        delimiter = (
            "\t"
            if path.suffix.lower() == ".tsv"
            else ","
        )

        encodings = (
            "utf-8-sig",
            "utf-8",
            "cp1252",
            "latin-1",
        )

        last_error: Exception | None = None

        for encoding in encodings:
            try:
                with path.open(
                    "r",
                    encoding=encoding,
                    newline="",
                ) as handle:
                    reader = csv.reader(
                        handle,
                        delimiter=delimiter,
                    )

                    output: list[str] = []
                    row_count = 0

                    for row in reader:
                        if row_count >= self.MAX_ROWS:
                            output.append(
                                "... [table truncated because "
                                "it exceeded the row limit] ..."
                            )
                            break

                        values = [
                            str(value)
                            for value in row[: self.MAX_COLUMNS]
                        ]

                        if not any(
                            value.strip()
                            for value in values
                        ):
                            continue

                        output.append(
                            " | ".join(values)
                        )

                        row_count += 1

                    return self._truncate(
                        "\n".join(output)
                    )

            except UnicodeDecodeError as exc:
                last_error = exc

        raise ValueError(
            f"Unable to decode spreadsheet '{path}'."
        ) from last_error

    def _read_spreadsheet(
        self,
        path: Path,
    ) -> str:
        extension = path.suffix.lower()

        if extension in EXCEL_EXTENSIONS:
            return self._read_excel(path)

        if extension in DELIMITED_EXTENSIONS:
            return self._read_delimited(path)

        if extension == ".xls":
            raise ValueError(
                "Legacy .xls files are not supported. "
                "Convert the file to .xlsx or .csv first."
            )

        raise ValueError(
            f"Unsupported spreadsheet format: {extension}"
        )

    def _truncate(
        self,
        text: str,
    ) -> str:
        if len(text) <= self.MAX_CHARS:
            return text

        return (
            text[: self.MAX_CHARS]
            + "\n\n"
            "[spreadsheet content truncated because "
            "it exceeded the character limit]"
        )

    def forward(
        self,
        file_path: str,
        question: str,
    ) -> str:
        try:
            if (
                not isinstance(file_path, str)
                or not file_path.strip()
            ):
                return (
                    "Error: file_path must be a "
                    "non-empty string."
                )

            if (
                not isinstance(question, str)
                or not question.strip()
            ):
                return (
                    "Error: question must be a "
                    "non-empty string."
                )

            if is_placeholder_path(file_path):
                return (
                    f"Error: File path '{file_path}' "
                    "is a placeholder or invalid."
                )

            path = resolve_file(
                self.base_dir,
                file_path,
            )

            if path is None:
                return (
                    f"Error: Spreadsheet '{file_path}' "
                    "was not found."
                )

            if not path.exists():
                return (
                    f"Error: Spreadsheet '{file_path}' "
                    "does not exist."
                )

            if not path.is_file():
                return (
                    f"Error: '{file_path}' is not a file."
                )

            spreadsheet_data = self._read_spreadsheet(path)

            if not spreadsheet_data.strip():
                return (
                    "Error: Spreadsheet contains no "
                    "readable data."
                )

            prompt = (
                "You are solving a GAIA benchmark task using "
                "spreadsheet data.\n\n"
                "Answer the user's question using ONLY the "
                "provided spreadsheet content.\n"
                "Do not invent values.\n"
                "Do not use outside knowledge when the answer "
                "must come from the spreadsheet.\n"
                "Perform arithmetic carefully when necessary.\n"
                "Return the precise answer requested.\n\n"
                f"User question:\n{question}\n\n"
                f"Spreadsheet:\n{spreadsheet_data}"
            )

            response = self.llm_service.generate_sync(
                messages=[
                    {
                        "role": "user",
                        "content": prompt,
                    }
                ],
                operation="llm.excel",
            )

            answer = str(response).strip()

            if not answer:
                return (
                    "Error: Excel analysis model returned "
                    "an empty response."
                )

            return answer

        except Exception as exc:
            return (
                "Error analyzing spreadsheet: "
                f"{type(exc).__name__}: {exc}"
            )


class ExcelTools:

    def __init__(
        self,
        llm_service: LLMService,
        base_dir: str = ".",
    ) -> None:
        if llm_service is None:
            raise ValueError(
                "ExcelTools requires an LLMService."
            )

        self.llm_service = llm_service
        self.base_dir = Path(base_dir).resolve()

    def get_tools(self) -> list[Tool]:
        return [
            AnalyzeExcelTool(
                llm_service=self.llm_service,
                base_dir=str(self.base_dir),
            )
        ]